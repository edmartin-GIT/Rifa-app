import sqlite3
from datetime import date
from io import BytesIO

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import openpyxl
from openpyxl.styles import Font

from db import init_db, get_conn, get_precio_tickera, set_precio_tickera

app = Flask(__name__)
app.secret_key = "rifa-control-secret"

MODALIDADES = ("CASH", "SQUARE", "ZELLE")

TICKET_BASE = 1000
TICKET_SIZE = 10
NUM_TICKERAS = 150


def tickera_a_rango(numero):
    inicio = TICKET_BASE + (numero - 1) * TICKET_SIZE
    fin = inicio + TICKET_SIZE - 1
    return inicio, fin


def ticket_inicio_a_tickera(ticket_inicio):
    if (ticket_inicio - TICKET_BASE) % TICKET_SIZE != 0:
        return None
    numero = (ticket_inicio - TICKET_BASE) // TICKET_SIZE + 1
    if 1 <= numero <= NUM_TICKERAS:
        return numero
    return None


def lista_tickeras(excluir_usadas=True, incluir_ademas=None):
    conn = get_conn()
    usadas = {r["tickera_numero"] for r in conn.execute("SELECT tickera_numero FROM transacciones")}
    conn.close()

    if not excluir_usadas:
        usadas = set()
    if incluir_ademas is not None:
        usadas.discard(incluir_ademas)

    return [
        {"numero": n, "ticket_inicio": tickera_a_rango(n)[0], "ticket_fin": tickera_a_rango(n)[1]}
        for n in range(1, NUM_TICKERAS + 1)
        if n not in usadas
    ]


def tickera_ya_vendida(numero, excluir_id=None):
    conn = get_conn()
    if excluir_id is None:
        row = conn.execute(
            "SELECT 1 FROM transacciones WHERE tickera_numero = ?", (numero,)
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT 1 FROM transacciones WHERE tickera_numero = ? AND id != ?", (numero, excluir_id)
        ).fetchone()
    conn.close()
    return row is not None


def calcular_totales(row):
    total_tickets = row["ticket_fin"] - row["ticket_inicio"] + 1
    total_esperado = row["precio_tickera"]
    saldo = round(total_esperado - row["monto_pagado"], 2)
    return total_tickets, total_esperado, saldo


def obtener_transacciones():
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM transacciones ORDER BY fecha_transaccion DESC, id DESC"
    ).fetchall()
    conn.close()
    return rows


def obtener_servidores_conocidos():
    conn = get_conn()
    rows = conn.execute(
        "SELECT DISTINCT nombre_servidor FROM transacciones ORDER BY nombre_servidor"
    ).fetchall()
    conn.close()
    return [r["nombre_servidor"] for r in rows]


def validar_formulario(form, excluir_id=None):
    errores = []
    nombre_servidor = form.get("nombre_servidor", "").strip()
    fecha_transaccion = form.get("fecha_transaccion", "").strip()
    modalidad_pago = form.get("modalidad_pago", "").strip().upper()
    numero_confirmacion = form.get("numero_confirmacion", "").strip()

    tickera_numero = None
    try:
        tickera_numero = int(form.get("tickera_numero", ""))
        if not (1 <= tickera_numero <= NUM_TICKERAS):
            errores.append(f"Tickera Numero debe estar entre 1 y {NUM_TICKERAS}.")
            tickera_numero = None
        elif tickera_ya_vendida(tickera_numero, excluir_id=excluir_id):
            errores.append(f"La Tickera {tickera_numero} ya fue vendida. Cada tickera es unica.")
            tickera_numero = None
    except ValueError:
        errores.append("Debe seleccionar una Tickera Numero valida.")

    ticket_inicio = ticket_fin = None
    if tickera_numero is not None:
        ticket_inicio, ticket_fin = tickera_a_rango(tickera_numero)

    try:
        monto_pagado = float(form.get("monto_pagado", "0") or 0)
    except ValueError:
        errores.append("El monto pagado debe ser un numero.")
        monto_pagado = 0

    if not nombre_servidor:
        errores.append("Nombre Servidor es requerido.")
    if not fecha_transaccion:
        errores.append("Fecha de Transaccion es requerida.")
    if modalidad_pago not in MODALIDADES:
        errores.append("Modalidad de Pago invalida.")
    if modalidad_pago in ("SQUARE", "ZELLE") and not numero_confirmacion:
        errores.append(f"Numero de Confirmacion es requerido para pagos {modalidad_pago}.")

    datos = {
        "tickera_numero": tickera_numero,
        "ticket_inicio": ticket_inicio,
        "ticket_fin": ticket_fin,
        "nombre_servidor": nombre_servidor,
        "fecha_transaccion": fecha_transaccion,
        "modalidad_pago": modalidad_pago,
        "numero_confirmacion": numero_confirmacion or None,
        "monto_pagado": monto_pagado,
    }
    return errores, datos


@app.route("/")
def index():
    precio_tickera = get_precio_tickera()
    transacciones = obtener_transacciones()

    filas = []
    total_por_modalidad = {m: 0.0 for m in MODALIDADES}
    total_por_tickera = {}
    saldo_total = 0.0
    total_recaudado = 0.0

    for t in transacciones:
        total_tickets, total_esperado, saldo = calcular_totales(t)
        filas.append(
            {
                **dict(t),
                "total_tickets": total_tickets,
                "total_esperado": total_esperado,
                "saldo": saldo,
            }
        )
        total_por_modalidad[t["modalidad_pago"]] += t["monto_pagado"]
        total_por_tickera.setdefault(t["tickera_numero"], 0.0)
        total_por_tickera[t["tickera_numero"]] += t["monto_pagado"]
        saldo_total += saldo
        total_recaudado += t["monto_pagado"]

    return render_template(
        "index.html",
        filas=filas,
        total_por_modalidad=total_por_modalidad,
        total_por_tickera=sorted(total_por_tickera.items()),
        saldo_total=round(saldo_total, 2),
        total_recaudado=round(total_recaudado, 2),
        precio_tickera=precio_tickera,
    )


@app.route("/nueva", methods=["GET", "POST"])
def nueva_transaccion():
    precio_tickera = get_precio_tickera()
    if request.method == "POST":
        errores, datos = validar_formulario(request.form)
        if errores:
            for e in errores:
                flash(e, "error")
            return render_template("form_transaccion.html", datos=datos, modalidades=MODALIDADES, titulo="Nueva Transaccion", precio_tickera=precio_tickera, tickeras=lista_tickeras(), servidores=obtener_servidores_conocidos())

        conn = get_conn()
        try:
            conn.execute(
                """INSERT INTO transacciones
                   (tickera_numero, ticket_inicio, ticket_fin, nombre_servidor, fecha_transaccion,
                    modalidad_pago, numero_confirmacion, precio_tickera, monto_pagado)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    datos["tickera_numero"],
                    datos["ticket_inicio"],
                    datos["ticket_fin"],
                    datos["nombre_servidor"],
                    datos["fecha_transaccion"],
                    datos["modalidad_pago"],
                    datos["numero_confirmacion"],
                    precio_tickera,
                    datos["monto_pagado"],
                ),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            flash(f"La Tickera {datos['tickera_numero']} ya fue vendida. Cada tickera es unica.", "error")
            return render_template("form_transaccion.html", datos=datos, modalidades=MODALIDADES, titulo="Nueva Transaccion", precio_tickera=precio_tickera, tickeras=lista_tickeras(), servidores=obtener_servidores_conocidos())
        conn.close()
        flash("Transaccion registrada correctamente.", "success")
        return redirect(url_for("index"))

    datos = {"fecha_transaccion": date.today().isoformat(), "monto_pagado": precio_tickera}
    return render_template("form_transaccion.html", datos=datos, modalidades=MODALIDADES, titulo="Nueva Transaccion", precio_tickera=precio_tickera, tickeras=lista_tickeras(), servidores=obtener_servidores_conocidos())


@app.route("/editar/<int:id>", methods=["GET", "POST"])
def editar_transaccion(id):
    conn = get_conn()
    transaccion = conn.execute("SELECT * FROM transacciones WHERE id = ?", (id,)).fetchone()
    conn.close()
    if transaccion is None:
        flash("Transaccion no encontrada.", "error")
        return redirect(url_for("index"))

    precio_tickera = transaccion["precio_tickera"]

    if request.method == "POST":
        errores, datos = validar_formulario(request.form, excluir_id=id)
        if errores:
            for e in errores:
                flash(e, "error")
            datos["id"] = id
            return render_template("form_transaccion.html", datos=datos, modalidades=MODALIDADES, titulo="Editar Transaccion", precio_tickera=precio_tickera, editar=True, tickeras=lista_tickeras(incluir_ademas=transaccion["tickera_numero"]), servidores=obtener_servidores_conocidos())

        conn = get_conn()
        conn.execute(
            """UPDATE transacciones SET
               tickera_numero=?, ticket_inicio=?, ticket_fin=?, nombre_servidor=?,
               fecha_transaccion=?, modalidad_pago=?, numero_confirmacion=?, monto_pagado=?
               WHERE id=?""",
            (
                datos["tickera_numero"],
                datos["ticket_inicio"],
                datos["ticket_fin"],
                datos["nombre_servidor"],
                datos["fecha_transaccion"],
                datos["modalidad_pago"],
                datos["numero_confirmacion"],
                datos["monto_pagado"],
                id,
            ),
        )
        conn.commit()
        conn.close()
        flash("Transaccion actualizada correctamente.", "success")
        return redirect(url_for("index"))

    datos = dict(transaccion)
    return render_template("form_transaccion.html", datos=datos, modalidades=MODALIDADES, titulo="Editar Transaccion", precio_tickera=precio_tickera, editar=True, tickeras=lista_tickeras(incluir_ademas=transaccion["tickera_numero"]), servidores=obtener_servidores_conocidos())


@app.route("/eliminar/<int:id>", methods=["POST"])
def eliminar_transaccion(id):
    conn = get_conn()
    conn.execute("DELETE FROM transacciones WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    flash("Transaccion eliminada.", "success")
    return redirect(url_for("index"))


@app.route("/configuracion", methods=["GET", "POST"])
def configuracion():
    if request.method == "POST":
        try:
            nuevo_precio = float(request.form.get("precio_tickera"))
            set_precio_tickera(nuevo_precio)
            flash("Precio por tickera actualizado. Aplica solo a transacciones nuevas.", "success")
        except (TypeError, ValueError):
            flash("Precio invalido.", "error")
        return redirect(url_for("configuracion"))

    return render_template("configuracion.html", precio_tickera=get_precio_tickera())


@app.route("/exportar")
def exportar_excel():
    transacciones = obtener_transacciones()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    encabezados = [
        "Tickera Numero", "Ticket Inicio", "Ticket Fin", "Total Tickets",
        "Nombre Servidor", "Fecha de Transaccion", "Modalidad de Pago",
        "Numero de Confirmacion", "Precio por Tickera", "Total Esperado",
        "Monto Pagado", "Saldo",
    ]
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    total_por_modalidad = {m: 0.0 for m in MODALIDADES}
    total_por_tickera = {}

    for t in transacciones:
        total_tickets, total_esperado, saldo = calcular_totales(t)
        ws.append([
            t["tickera_numero"], t["ticket_inicio"], t["ticket_fin"], total_tickets,
            t["nombre_servidor"], t["fecha_transaccion"], t["modalidad_pago"],
            t["numero_confirmacion"] or "", t["precio_tickera"], total_esperado,
            t["monto_pagado"], saldo,
        ])
        total_por_modalidad[t["modalidad_pago"]] += t["monto_pagado"]
        total_por_tickera.setdefault(t["tickera_numero"], 0.0)
        total_por_tickera[t["tickera_numero"]] += t["monto_pagado"]

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    ws_modalidad = wb.create_sheet("Total por Modalidad")
    ws_modalidad.append(["Modalidad de Pago", "Total"])
    for cell in ws_modalidad[1]:
        cell.font = Font(bold=True)
    for modalidad, total in total_por_modalidad.items():
        ws_modalidad.append([modalidad, total])

    ws_tickera = wb.create_sheet("Total por Tickera")
    ws_tickera.append(["Tickera Numero", "Total Acumulado"])
    for cell in ws_tickera[1]:
        cell.font = Font(bold=True)
    for tickera, total in sorted(total_por_tickera.items()):
        ws_tickera.append([tickera, total])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"control_rifa_{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


init_db()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
